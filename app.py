import gradio as gr
import pandas as pd
import json, os
from datetime import datetime
from collections import Counter

# data pengganti nama fasilitas
replacement_nama_fasilitas = {
    "AEON Credit Services Indonesia": "AEON Credit",
    "Adira Dinamika Multi Finance": "Adira",
    "Akulaku Finance Indonesia": "Akulaku",
    "Atome Finance Indonesia": "Atome Finance",
    "Astra Multi Finance": "Astra MF",
    "BFI Finance Indonesia": "BFI",
    "BIMA Multi Finance": "Bima MF",
    "BPD Jawa Barat dan Banten": "BJB",
    "BPD Jawa Barat dan Banten Syariah": "BJB Syariah",
    "BPD Jawa Timur": "Bank Jatim",
    "BPD Sumatera Utara": "Bank Sumut",
    "Bank BCA Syariah": "BCA Syariah",
    "Bank CIMB Niaga": "CIMB Niaga",
    "Bank Central Asia": "BCA",
    "Bank DBS Indonesia": "Bank DBS",
    "Bank Danamon": "Danamon",
    "Bank Danamon Indonesia": "Danamon",
    "Bank Danamon Syariah": "Danamon Syariah",
    "Bank Hibank Indonesia": "Hibank",
    "Bank HSBC Indonesia": "HSBC",
    "Bank KEB Hana Indonesia": "Bank KEB Hana",
    "Bank Mandiri": "Bank Mandiri",
    "Bank Mandiri Taspen": "Bank Mantap",
    "Bank Mayapada Internasional": "Bank Mayapada",
    "Bank Maybank Indonesia": "Maybank",
    "Bank Mega Syariah": "Bank Mega Syariah",
    "Bank Muamalat Indonesia": "Bank Muamalat",
    "Bank Negara Indonesia": "BNI",
    "Bank Neo Commerce": "Akulaku",
    "Bank OCBC NISP": "OCBC NISP",
    "Bank Panin Indonesia": "Panin Bank",
    "Bank Permata": "Bank Permata",
    "Bank Rakyat Indonesia": "BRI",
    "Bank Sahabat Sampoerna": "Bank Sampoerna",
    "Bank Saqu Indonesia ( ": "Bank Saqu",
    "Bank Seabank Indonesia": "Seabank",
    "Bank SMBC Indonesia": "Bank SMBC",
    "Bank Syariah Indonesia": "BSI",
    "Bank Tabungan Negara": "BTN",
    "Bank UOB Indonesia": "Bank UOB",
    "Bank Woori Saudara": "BWS",
    "Bank Woori Saudara Indonesia 1906":"BWS",
    "Bussan Auto Finance": "BAF",
    "Cakrawala Citra Mega Multifinance":"CCM Finance",
    "Commerce Finance": "Seabank",
    "Dana Mandiri Sejahtera": "Dana Mandiri",
    "Esta Dana Ventura": "Esta Dana",
    "Federal International Finance": "FIF",
    "Home Credit Indonesia": "Home Credit",
    "Indodana Multi Finance": "Indodana MF",
    "Indomobil Finance Indonesia": "IMFI",
    "Indonesia Airawata Finance (": "Indonesia Airawata Finance",
    "JACCS Mitra Pinasthika Mustika Finance Indonesia": "JACCS",
    "KB Finansia Multi Finance": "Kreditplus",
    "Kredivo Finance Indonesia": "Kredivo",
    "Krom Bank Indonesia": "Krom Bank",
    "Mandala Multifinance": "Mandala MF",
    "Mandiri Utama Finance": "MUF",
    "Maybank Syariah": "Maybank Syariah",
    "Mega Auto Finance": "MAF",
    "Mega Central Finance": "MCF",
    "Mitra Bisnis Keluarga Ventura": "MBK",
    "Multifinance Anak Bangsa": "MF Anak Bangsa",
    "Panin Bank": "Panin Bank",
    "Permodalan Nasional Madani": "PNM",
    "Pratama Interdana Finance": "Pratama Finance",
    "Standard Chartered Bank": "Standard Chartered",
    "Summit Oto Finance": "Summit Oto",
    "Super Bank Indonesia": "Superbank",
    "Wahana Ottomitra Multiartha": "WOM",
    "Bank Jago": "Bank Jago",
    "Bank BTPN Syariah,": "BTPNS",
    "Bina Artha Ventura": "BAV"
}

# pembersihan nama fasilitas
def bersihkan_nama_fasilitas(nama_fasilitas: str) -> str:
    if not nama_fasilitas:
        return ""
    lower_fasilitas = nama_fasilitas.lower()
    if "d/h" in lower_fasilitas:
        nama_bersih = nama_fasilitas[:lower_fasilitas.find("d/h")].strip()
    elif "d.h" in lower_fasilitas:
        nama_bersih = nama_fasilitas[:lower_fasilitas.find("d.h")].strip()
    else:
        nama_bersih = nama_fasilitas.strip()
    for pattern in ["PT ", "PT.","PD.", "(Persero)", "(Perseroda)", "(UUS)", " Tbk"]:
        nama_bersih = nama_bersih.replace(pattern, "")
        nama_bersih = nama_bersih.replace("Bank Perekonomian Rakyat Syariah", "BPRS")
    nama_bersih = nama_bersih.replace("Bank Perekonomian Rakyat", "BPR")
    nama_bersih = nama_bersih.replace("Koperasi Simpan Pinjam", "KSP")
    nama_bersih = nama_bersih.strip()
    for nama_asli, alias in replacement_nama_fasilitas.items():
        if nama_asli.lower() == nama_bersih.lower():
            return alias
    return nama_bersih

# gabung banyak data menjadi satu data
def gabungkan_fasilitas_dengan_jumlah(fasilitas_list):
    counter = Counter(fasilitas_list)
    return '; '.join([f"{nama} ({jumlah})" if jumlah > 1 else nama for nama, jumlah in counter.items()])

# proses semua file
def proses_files_gradio(files):
    """
    files: list of gradio.FileData
    Return: (pandas.DataFrame, path_excel)
    """
    if not files:
        return pd.DataFrame(), None

    hasil_semua = []

    for f in files:
        # nama asli file (untuk NIK)
        original_name = getattr(f, "orig_name", None) or getattr(f, "name", None) or os.path.basename(f.name if hasattr(f, "name") else f)
        path = getattr(f, "name", None) or getattr(f, "path", None) or f

        if not str(original_name).lower().endswith(".txt"):
            # lewati file non-txt
            continue

        try:
            with open(path, "r", encoding="latin-1") as file:
                data = json.load(file)
        except Exception as e:
            # lewati file bermasalah 
            print(f"Gagal membaca file: {original_name} -> {e}")
            continue

        # ambil data perfasilitas
        fasilitas = data.get('individual', {}).get('fasilitas', {}).get('kreditPembiayan', [])
        data_pokok = data.get('individual', {}).get('dataPokokDebitur', [])
        nama_debitur = ', '.join(set(debitur.get('namaDebitur', '') for debitur in data_pokok if debitur.get('namaDebitur')))

        total_plafon = 0
        total_baki_debet = 0
        jumlah_fasilitas_aktif = 0
        kol_1_list, kol_25_list, wo_list, lovi_list = [], [], [], []
        baki_debet_kol25wo = 0
        excluded_fasilitas = {"BTPNS", "Bank Jago", "BAV"}

        for item in fasilitas:
            kondisi_ket = (item.get('kondisiKet') or '').lower()
            nama_fasilitas = item.get('ljkKet') or ''
            nama_fasilitas_lower = nama_fasilitas.lower()

            # pemampilan kodisi keterangan lunas hanya untuk lovi
            if kondisi_ket == "lunas" and "pt lolc ventura indonesia" not in nama_fasilitas_lower:
                continue

            jumlah_hari_tunggakan = int(item.get('jumlahHariTunggakan', 0))
            kualitas = item.get('kualitas', '')
            kol_value = f"{kualitas}/{jumlah_hari_tunggakan}" if jumlah_hari_tunggakan != 0 else kualitas
            tanggal_kondisi = item.get('tanggalKondisi', '')
            baki_debet = int(item.get('bakiDebet', 0))

            if kondisi_ket in ['dihapusbukukan', 'hapus tagih', 'fasilitas aktif'] and baki_debet == 0:
                baki_debet = sum([
                    int(item.get('tunggakanPokok', 0)),
                    int(item.get('tunggakanBunga', 0)),
                    int(item.get('denda', 0))
                ])
                if baki_debet == 0:
                    kondisi_ket = 'lunas'

            plafon_awal = int(item.get('plafonAwal', 0))
            nama_fasilitas_bersih = bersihkan_nama_fasilitas(nama_fasilitas)
            baki_debet_format = "{:,.0f}".format(baki_debet).replace(",", ".")

            if kondisi_ket == 'fasilitas aktif' and kualitas == '1' and jumlah_hari_tunggakan <= 30:
                fasilitas_teks = nama_fasilitas_bersih
            elif kondisi_ket == 'fasilitas aktif':
                fasilitas_teks = f"{nama_fasilitas_bersih} Kol {kol_value} {baki_debet_format}"
            elif kondisi_ket in ['dihapusbukukan', 'hapus tagih']:
                try:
                    tahun_wo = int(str(tanggal_kondisi)[:4])
                except:
                    tahun_wo = ""
                fasilitas_teks = f"{nama_fasilitas_bersih} WO {tahun_wo} {baki_debet_format}"
            else:
                fasilitas_teks = nama_fasilitas_bersih

            # ambil satu data dengan tanggalAkadAkhir terbaru khusus kolom lovi
            if kondisi_ket == 'lunas':
                fasilitas_lovi = "Lunas"
            elif kondisi_ket == 'fasilitas aktif':
                fasilitas_lovi = f"Kol {kol_value}"
            elif kondisi_ket in ['dihapusbukukan', 'hapus tagih']:
                fasilitas_lovi = f"WO {tahun_wo} {baki_debet_format}"
            else:
                fasilitas_lovi = nama_fasilitas_bersih

            if "pt lolc ventura indonesia" not in nama_fasilitas_lower:
                if kondisi_ket == "fasilitas aktif":
                    total_plafon += plafon_awal
                    total_baki_debet += baki_debet
                    jumlah_fasilitas_aktif += 1
                    if kondisi_ket == 'fasilitas aktif' and kualitas == '1' and jumlah_hari_tunggakan <= 30:
                        if jumlah_hari_tunggakan == 0:
                            kol_1_list.append(nama_fasilitas_bersih)
                        else:
                            kol_1_list.append(f"{nama_fasilitas_bersih} Kol {kualitas}/{jumlah_hari_tunggakan}")
                    else:
                        kol_25_list.append(fasilitas_teks)
                        if nama_fasilitas_bersih not in excluded_fasilitas:
                            baki_debet_kol25wo += baki_debet
                elif kondisi_ket in ['dihapusbukukan', 'hapus tagih']:
                    wo_list.append(fasilitas_teks)
                    if nama_fasilitas_bersih not in excluded_fasilitas:
                        baki_debet_kol25wo += baki_debet
            else:
                if kondisi_ket in ["fasilitas aktif", "lunas", "dihapusbukukan"]:
                    tanggal_akad_akhir = item.get("tanggalAkadAkhir", "")
                    if tanggal_akad_akhir:
                        if not lovi_list:
                            lovi_list.append({
                                "keterangan": fasilitas_lovi,
                                "tanggal": tanggal_akad_akhir
                            })
                        else:
                            if tanggal_akad_akhir > lovi_list[0]["tanggal"]:
                                lovi_list[0] = {
                                    "keterangan": fasilitas_lovi,
                                    "tanggal": tanggal_akad_akhir
                                }

        # logika pemberian rekomendasi
        if (
            jumlah_fasilitas_aktif >= 0 and
            not kol_25_list and
            not wo_list and
            not lovi_list
        ):
            rekomendasi = "OK"
        elif any("lunas" in lovi.get('keterangan', '').lower() or "kol 1" in lovi.get('keterangan', '').lower() for lovi in lovi_list):
            rekomendasi = "OK"
        elif (
            jumlah_fasilitas_aktif >= 0 and
            baki_debet_kol25wo <= 250_000 and
            not lovi_list
        ):
            rekomendasi = "OK"
        else:
            rekomendasi = "NOT OK"

        # mengambil data NIK dari nama file
        filename = os.path.basename(original_name or path)
        nik = os.path.splitext(filename)[0]
        if nik.upper().startswith("NIK_"):
            nik = nik[4:]

        hasil_semua.append({
            'NIK': "'" + nik,
            'Nama Debitur': nama_debitur,
            'Rekomendasi': rekomendasi,
            'Jumlah Fasilitas': jumlah_fasilitas_aktif,
            'Total Plafon Awal': total_plafon if jumlah_fasilitas_aktif > 0 else "",
            'Total Baki Debet': total_baki_debet if jumlah_fasilitas_aktif > 0 else "",
            'Kol 1': gabungkan_fasilitas_dengan_jumlah(kol_1_list),
            'Kol 2-5': '; '.join(kol_25_list),
            'WO/dihapusbukukan': '; '.join(wo_list),
            'LOVI': '; '.join([l.get('keterangan', '') for l in lovi_list])
        })

    # jika tidak ada data valid
    if not hasil_semua:
        return pd.DataFrame(), None

    # simpan ke excel 
    df = pd.DataFrame(hasil_semua)

    tanggal_hari_ini = datetime.today().strftime('%d-%m-%Y_%H%M%S')
    output_file = f'File SLIK {tanggal_hari_ini}.xlsx'
    df.to_excel(output_file, index=False)

    # format excel (openpyxl)
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    #custom ukuran tabel di excel
    custom_widths = {
        'NIK': 17,
        'Nama Debitur': 22,
        'Rekomendasi': 12,
        'Jumlah Fasilitas': 8,
        'Total Plafon Awal': 13,
        'Total Baki Debet': 13,
        'Kol 1': 31,
        'Kol 2-5': 31,
        'WO/dihapusbukukan': 31,
        'LOVI': 9
    }

    #custom format kolom di excel
    wrap_columns = {'Kol 1', 'Kol 2-5', 'WO/dihapusbukukan', 'LOVI'}
    center_columns = {'NIK', 'Rekomendasi', 'Jumlah Fasilitas', 'Kol 1', 'Kol 2-5', 'WO/dihapusbukukan', 'LOVI'}
    number_format_columns = {'Total Plafon Awal', 'Total Baki Debet'}

    # all border di excel
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_row = ws[1]
    header = [cell.value for cell in header_row]

    for idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(idx)
        col_name = header[idx - 1] if idx - 1 < len(header) else ''

        wrap = col_name in wrap_columns
        center = col_name in center_columns

        if center and wrap:
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif center:
            alignment = Alignment(horizontal='center', vertical='center')
        elif wrap:
            alignment = Alignment(wrap_text=True)
        else:
            alignment = Alignment()

        for i, cell in enumerate(col_cells):
            cell.alignment = alignment
            cell.font = Font(size=8)
            cell.border = thin_border
            if i != 0 and col_name in number_format_columns:
                cell.number_format = '#,##0'

        if col_name in custom_widths:
            ws.column_dimensions[col_letter].width = custom_widths[col_name]
        else:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file)

    # Kembalikan DataFrame & path file untuk diunduh
    return df, output_file

# ui gradio

# button clear data
def clear_data():
    return None, None, pd.DataFrame()

with gr.Blocks(theme=gr.themes.Soft(), css="""
#preview-table { overflow-x: auto; }
#preview-table table { font-size: 12px !important; white-space: nowrap !important; width: 100%; }
#preview-table thead th { background: #0074D9 !important; color: white !important; text-align: center !important; padding: 6px !important; }
footer {text-align:center; margin-top:10px; color:gray; font-size:13px}
""") as demo:

    gr.Markdown("<h1 style='text-align:center'>📄 Sistem Proses Data Debitur</h1>")
    gr.Markdown(
        "Unggah beberapa file <code>.txt</code>, kemudian klik <b>Proses</b>, lalu unduh file Excel hasil pengolahan. "
    )

    # row untuk upload dan download
    with gr.Row():
        with gr.Column(scale=1):
            inp_files = gr.File(label="📤 Upload file .txt", file_count="multiple", file_types=[".txt"])
            note = gr.Markdown("Tip: Jika tidak ada output, pastikan file berekstensi <code>.txt</code>.")
            btn = gr.Button("🚀 Proses", variant="primary")
        with gr.Column(scale=1):
            out_file = gr.File(label="⬇️ Unduh File Excel", file_types=[".xlsx"])
            note = gr.Markdown("Note: Klik ukuran file disebelah kanan nama file untuk mengunduh.")
            clear_btn = gr.Button("🗑️ Clear Data", variant="secondary")

    # Preview tabel
    out_df = gr.Dataframe(label="Preview Hasil Proses Data Debitur", elem_id="preview-table", wrap=False)

    btn.click(fn=proses_files_gradio, inputs=[inp_files], outputs=[out_df, out_file])
    clear_btn.click(fn=clear_data, outputs=[inp_files, out_file, out_df])

    gr.HTML("<footer>© 2025 - Sistem Proses Data Debitur | Dibuat dengan oleh Ayu Nurhasanah</footer>")

demo.launch()

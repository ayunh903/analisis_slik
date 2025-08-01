import json
import os
import tempfile
from datetime import datetime
from collections import Counter
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from flask import Flask, request, send_file
import pandas as pd

app = Flask(__name__)

replacement_nama_fasilitas = {
    "AEON Credit Services Indonesia": "AEON Credit",
    "Adira Dinamika Multi Finance": "Adira",
    "Akulaku Finance Indonesia": "Akulaku",
    "Atome Finance Indonesia": "Atome Finance",
    "Astra Multi Finance": "Astra MF",
    "BFI Finance Indonesia": "BFI",
    "BIMA Multi Finance": "Bima MF",
    "BPD Jawa Barat dan Banten": "BJB",
    "BPD Jawa Barat dan Banten Syariah": "BJB Syariah",
    "BPD Jawa Timur": "Bank Jatim",
    "BPD Sumatera Utara": "Bank Sumut",
    "Bank BCA Syariah": "BCA Syariah",
    "Bank CIMB Niaga": "CIMB Niaga",
    "Bank Central Asia": "BCA",
    "Bank DBS Indonesia": "Bank DBS",
    "Bank Danamon": "Danamon",
    "Bank Danamon Indonesia": "Danamon",
    "Bank Danamon Syariah": "Danamon Syariah",
    "Bank Hibank Indonesia": "Hibank",
    "Bank HSBC Indonesia": "HSBC",
    "Bank KEB Hana Indonesia": "Bank KEB Hana",
    "Bank Mandiri": "Bank Mandiri",
    "Bank Mandiri Taspen": "Bank Mantap",
    "Bank Mayapada Internasional": "Bank Mayapada",
    "Bank Maybank Indonesia": "Maybank",
    "Bank Mega Syariah": "Bank Mega Syariah",
    "Bank Muamalat Indonesia": "Bank Muamalat",
    "Bank Negara Indonesia": "BNI",
    "Bank Neo Commerce": "Akulaku",
    "Bank OCBC NISP": "OCBC NISP",
    "Bank Panin Indonesia": "Panin Bank",
    "Bank Permata": "Bank Permata",
    "Bank Rakyat Indonesia": "BRI",
    "Bank Sahabat Sampoerna": "Bank Sampoerna",
    "Bank Saqu Indonesia ( ": "Bank Saqu",
    "Bank Seabank Indonesia": "Seabank",
    "Bank SMBC Indonesia": "Bank SMBC",
    "Bank Syariah Indonesia": "BSI",
    "Bank Tabungan Negara": "BTN",
    "Bank UOB Indonesia": "Bank UOB",
    "Bank Woori Saudara": "BWS",
    "Bank Woori Saudara Indonesia 1906":"BWS",
    "Bussan Auto Finance": "BAF",
    "Cakrawala Citra Mega Multifinance":"CCM Finance",
    "Commerce Finance": "Seabank",
    "Dana Mandiri Sejahtera": "Dana Mandiri",
    "Esta Dana Ventura": "Esta Dana",
    "Federal International Finance": "FIF",
    "Home Credit Indonesia": "Home Credit",
    "Indodana Multi Finance": "Indodana MF",
    "Indomobil Finance Indonesia": "IMFI",
    "Indonesia Airawata Finance (": "Indonesia Airawata Finance",
    "JACCS Mitra Pinasthika Mustika Finance Indonesia": "JACCS",
    "KB Finansia Multi Finance": "Kreditplus",
    "Kredivo Finance Indonesia": "Kredivo",
    "Krom Bank Indonesia": "Krom Bank",
    "Mandala Multifinance": "Mandala MF",
    "Mandiri Utama Finance": "MUF",
    "Maybank Syariah": "Maybank Syariah",
    "Mega Auto Finance": "MAF",
    "Mega Central Finance": "MCF",
    "Mitra Bisnis Keluarga Ventura": "MBK",
    "Multifinance Anak Bangsa": "MF Anak Bangsa",
    "Panin Bank": "Panin Bank",
    "Permodalan Nasional Madani": "PNM",
    "Pratama Interdana Finance": "Pratama Finance",
    "Standard Chartered Bank": "Standard Chartered",
    "Summit Oto Finance": "Summit Oto",
    "Super Bank Indonesia": "Superbank",
    "Wahana Ottomitra Multiartha": "WOM",
    "Bank Jago": "Bank Jago",
    "Bank BTPN Syariah,": "BTPNS",
    "Bina Artha Ventura": "BAV"
}


def bersihkan_nama_fasilitas(nama):
    if not nama:
        return ""
    nama = nama.strip().replace("PT ", "").replace("PT.", "")
    return replacement_nama_fasilitas.get(nama, nama)

def gabungkan_fasilitas_dengan_jumlah(fasilitas_list):
    counter = Counter(fasilitas_list)
    return '; '.join([f"{nama} ({jumlah})" if jumlah > 1 else nama for nama, jumlah in counter.items()])

@app.route("/api/proses", methods=["POST"])
def proses():
    files = request.files.getlist("files")
    hasil_semua = []

    for file in files:
        try:
            content = file.read().decode("latin-1")
            data = json.loads(content)
        except:
            continue

        fasilitas = data.get('individual', {}).get('fasilitas', {}).get('kreditPembiayan', [])
        data_pokok = data.get('individual', {}).get('dataPokokDebitur', [])
        nama_debitur = ', '.join(set(d.get('namaDebitur', '') for d in data_pokok if d.get('namaDebitur')))

        total_plafon = 0
        total_baki_debet = 0
        jumlah_fasilitas_aktif = 0
        kol_1_list, kol_25_list, wo_list, lovi_list = [], [], [], []
        baki_debet_kol25wo = 0

        for item in fasilitas:
            kondisi_ket = (item.get('kondisiKet') or '').lower()
            nama_fasilitas = item.get('ljkKet') or ''
            nama_fasilitas_bersih = bersihkan_nama_fasilitas(nama_fasilitas)

            jumlah_hari_tunggakan = int(item.get('jumlahHariTunggakan', 0))
            kualitas = item.get('kualitas', '')
            kol_value = f"{kualitas}/{jumlah_hari_tunggakan}" if jumlah_hari_tunggakan != 0 else kualitas
            tanggal_kondisi = item.get('tanggalKondisi', '')
            baki_debet = int(item.get('bakiDebet', 0))

            if kondisi_ket in ['dihapusbukukan', 'hapus tagih', 'fasilitas aktif'] and baki_debet == 0:
                baki_debet = sum([
                    int(item.get('tunggakanPokok', 0)),
                    int(item.get('tunggakanBunga', 0)),
                    int(item.get('denda', 0))
                ])
                if baki_debet == 0:
                    kondisi_ket = 'lunas'

            plafon_awal = int(item.get('plafonAwal', 0))
            baki_debet_format = "{:,.0f}".format(baki_debet).replace(",", ".")

            if kondisi_ket == 'fasilitas aktif' and kualitas == '1' and jumlah_hari_tunggakan <= 30:
                kol_1_list.append(nama_fasilitas_bersih)
            elif kondisi_ket == 'fasilitas aktif':
                kol_25_list.append(f"{nama_fasilitas_bersih} Kol {kol_value} {baki_debet_format}")
                baki_debet_kol25wo += baki_debet
            elif kondisi_ket in ['dihapusbukukan', 'hapus tagih']:
                try:
                    tahun_wo = int(str(tanggal_kondisi)[:4])
                except:
                    tahun_wo = ""
                wo_list.append(f"{nama_fasilitas_bersih} WO {tahun_wo} {baki_debet_format}")
                baki_debet_kol25wo += baki_debet

            if kondisi_ket == 'fasilitas aktif':
                total_plafon += plafon_awal
                total_baki_debet += baki_debet
                jumlah_fasilitas_aktif += 1

        rekomendasi = "OK" if not kol_25_list and not wo_list and not lovi_list else "NOT OK"

        filename = file.filename or "file.txt"
        nik = os.path.splitext(filename)[0].replace("NIK_", "")

        hasil_semua.append({
            'NIK': "'" + nik,
            'Nama Debitur': nama_debitur,
            'Rekomendasi': rekomendasi,
            'Jumlah Fasilitas': jumlah_fasilitas_aktif,
            'Total Plafon Awal': total_plafon,
            'Total Baki Debet': total_baki_debet,
            'Kol 1': gabungkan_fasilitas_dengan_jumlah(kol_1_list),
            'Kol 2-5': '; '.join(kol_25_list),
            'WO/dihapusbukukan': '; '.join(wo_list),
            'LOVI': '; '.join([l.get('keterangan', '') for l in lovi_list])
        })

    if not hasil_semua:
        return {"error": "Tidak ada file valid"}, 400

    df = pd.DataFrame(hasil_semua)

    tempdir = tempfile.gettempdir()
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_path = os.path.join(tempdir, f"Hasil_SLIK_{timestamp}.xlsx")
    df.to_excel(output_path, index=False)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    wrap_columns = {'Kol 1', 'Kol 2-5', 'WO/dihapusbukukan', 'LOVI'}
    center_columns = {'NIK', 'Rekomendasi', 'Jumlah Fasilitas', 'Kol 1', 'Kol 2-5', 'WO/dihapusbukukan', 'LOVI'}
    number_format_columns = {'Total Plafon Awal', 'Total Baki Debet'}

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(idx)
        col_name = header[idx - 1] if idx - 1 < len(header) else ''

        wrap = col_name in wrap_columns
        center = col_name in center_columns

        if center and wrap:
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif center:
            alignment = Alignment(horizontal='center', vertical='center')
        elif wrap:
            alignment = Alignment(wrap_text=True)
        else:
            alignment = Alignment()

        for i, cell in enumerate(col_cells):
            cell.alignment = alignment
            cell.font = Font(size=8)
            cell.border = thin_border
            if i != 0 and col_name in number_format_columns:
                cell.number_format = '#,##0'

    wb.save(output_path)

    return send_file(output_path, as_attachment=True, download_name="hasil_slik.xlsx")
